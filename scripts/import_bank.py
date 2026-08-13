#!/usr/bin/env python3
"""문제 은행 엑셀에서 게임 데이터(TypeScript)를 생성한다.

    python3 scripts/import_bank.py [엑셀경로]

엑셀이 원본이다. shared/src/data/*.ts 는 생성물이므로 직접 고치지 말 것.

게임에 넣는 기준은 두 가지다.
  - 분류체계 시트의 '게임 노출' 이 '노출' 인 분류만 사용한다.
  - 검수상태가 '검수완료' 인 문항만 사용한다.

보류 문항을 빼는 이유: 이 게임은 오답일 때 해설을 그대로 화면에 띄워
참가자를 학습시킨다. 출처에서 확인되지 않은 내용을 가르치면 안 된다.
엑셀에는 남아 있으므로, 확인이 끝나 검수완료로 바뀌면 자동으로 합류한다.
"""

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'docs' / '문제은행_v4.xlsx'
DATA = ROOT / 'shared' / 'src' / 'data'

MAX_TEXT, MAX_EXPL = 40, 45

# 정답 분류 → (게임 내부 ID, 버튼에 쓸 짧은 이름)
IDS = {
    '데이터관리센터': ('data-center', '데이터관리'),
    '특허정보검색서비스(KIPRIS)': ('kipris', 'KIPRIS 검색'),
    '특허문서전자화센터': ('digitization', '문서 전자화'),
    '한국특허영문초록(KPA)': ('kpa', '영문초록 KPA'),
    '특허정보 번역서비스': ('translation', '번역서비스'),
    '지식재산 통계서비스': ('statistics', '지식재산 통계'),
    'IP정보통합센터': ('ipic', 'IP정보통합'),
    '특허정보활용서비스(KIPRISPlus)': ('kipris-plus', 'KIPRISPlus'),
    '국제특허정보박람회(PATINEX)': ('patinex', 'PATINEX'),
    '지식재산 정보화 R&D': ('rnd', '정보화 R&D'),
    '특허넷시스템 운영': ('kiponet', '특허넷 운영'),
    '특허고객 상담서비스': ('customer', '고객상담'),
    '지식재산 정보화 국제협력': ('global', '국제협력'),
    '지식재산 기술지원': ('tech-transfer', '기술지원'),
}


def _one_line(v, limit: int = 34) -> str:
    """주요 업무 설명을 버튼에 들어갈 한 줄로 줄인다."""
    s = str(v or '').split('.')[0].strip()
    return s if len(s) <= limit else s[: limit - 1].rstrip(' ,·') + '…'


def ts(s: str) -> str:
    return "'" + str(s).replace('\\', '\\\\').replace("'", "\\'") + "'"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    cat_ws, q_ws = wb['분류체계'], wb['문제은행']

    # ── 분류체계 ────────────────────────────────────────────────────
    categories = []
    for r in range(2, cat_ws.max_row + 1):
        name = cat_ws.cell(r, 1).value
        if not name or name == '합계':
            continue
        exposed = str(cat_ws.cell(r, 7).value or '').strip() == '노출'
        if not exposed:
            print(f'  분류 제외: {name} (게임 노출 아님)')
            continue
        if name not in IDS:
            raise SystemExit(f'ID 미등록 분류: {name} — IDS에 추가하세요')
        unit = str(cat_ws.cell(r, 4).value or '').strip()
        # 버튼 배지는 실 단위만 쓴다. 본부명은 5개 사업이 공유해 변별이 안 되고,
        # 버튼 폭에서 줄바꿈이 단어 중간에 걸린다.
        unit_short = unit.split('·')[-1].strip() if unit else ''
        categories.append({
            'name': name,
            'area': cat_ws.cell(r, 2).value,
            'unit': unit,
            'unitShort': unit_short,
            'desc': _one_line(cat_ws.cell(r, 3).value),
            'id': IDS[name][0],
            'short': IDS[name][1],
        })
    exposed_names = {c['name'] for c in categories}

    # ── 문항 ────────────────────────────────────────────────────────
    rows, skipped_review, skipped_cat = [], 0, 0
    for r in range(2, q_ws.max_row + 1):
        qid = q_ws.cell(r, 1).value
        if not qid:
            continue
        correct = q_ws.cell(r, 3).value
        if str(q_ws.cell(r, 12).value or '').strip() != '검수완료':
            skipped_review += 1
            continue
        if correct not in exposed_names:
            skipped_cat += 1
            continue
        # 게임에서 빠진 분류는 오답 후보에서도 빼야 버튼에 안 뜬다
        distractors = [q_ws.cell(r, c).value for c in (5, 6, 7)]
        distractors = [d for d in distractors if d and d in exposed_names and d != correct]
        rows.append({
            'id': qid,
            'text': str(q_ws.cell(r, 2).value).strip(),
            'correct': correct,
            'distractors': distractors,
            'explanation': str(q_ws.cell(r, 8).value).strip(),
            'difficulty': int(q_ws.cell(r, 9).value),
            'source': str(q_ws.cell(r, 11).value or '').strip(),
        })

    # ── 검증 ────────────────────────────────────────────────────────
    problems = []
    seen_ids = set()
    for q in rows:
        if q['id'] in seen_ids:
            problems.append(f"{q['id']}: 문항ID 중복")
        seen_ids.add(q['id'])
        if len(q['text']) > MAX_TEXT:
            problems.append(f"{q['id']}: 민원 내용 {len(q['text'])}자 ({MAX_TEXT}자 초과)")
        if len(q['explanation']) > MAX_EXPL:
            problems.append(f"{q['id']}: 해설 {len(q['explanation'])}자 ({MAX_EXPL}자 초과)")
        if q['difficulty'] not in (1, 2, 3):
            problems.append(f"{q['id']}: 난이도 {q['difficulty']}")
        if not q['source']:
            problems.append(f"{q['id']}: 출처 없음")
    if problems:
        raise SystemExit('데이터 오류:\n  ' + '\n  '.join(problems))

    # ── departments.ts ──────────────────────────────────────────────
    out = ['import type { Department } from \'../rules.js\';', '',
           '/**', ' * 정답 분류 — www.kipi.or.kr 주요사업 기준.',
           ' *', f' * docs/{XLSX.name} 에서 생성됩니다. 직접 고치지 마세요.',
           ' */', 'export const DEPARTMENTS: Department[] = [']
    for c in categories:
        out.append(f"  {{ id: {ts(c['id'])}, name: {ts(c['name'])}, short: {ts(c['short'])},")
        out.append(f"    unit: {ts(c['unit'])}, unitShort: {ts(c['unitShort'])}, desc: {ts(c['desc'])} }},")
    out += ['];', '',
            'export const DEPARTMENT_BY_ID = new Map(DEPARTMENTS.map((d) => [d.id, d]));', '',
            '/** 키보드 단축키 — 페이즈 3의 12개 버튼까지 커버한다. */',
            "export const HOTKEYS = ['Q', 'W', 'E', 'R', 'A', 'S', 'D', 'F', 'Z', 'X', 'C', 'V'];", '']
    (DATA / 'departments.ts').write_text('\n'.join(out), encoding='utf-8')

    # ── complaints.ts ───────────────────────────────────────────────
    # 출처 URL이 문항마다 반복되면 파일과 번들이 함께 커진다. 표로 빼서
    # 인덱스로 참조하면 같은 데이터가 절반 크기로 줄어든다.
    sources = []
    for q in rows:
        if q['source'] not in sources:
            sources.append(q['source'])

    out = ["import type { Complaint } from '../rules.js';", '',
           '/**', f' * 검수완료 문항 {len(rows)}개 — www.kipi.or.kr 주요사업 기준 (조사일 2026-08-13).',
           ' *', f' * docs/{XLSX.name} 에서 생성됩니다. 문항 추가·수정은 엑셀에서 하고',
           ' * scripts/import_bank.py 를 다시 실행하세요.',
           ' */', '', '/** 출처 URL 표. 문항은 인덱스로만 참조한다. */',
           'const SOURCES = [']
    for s in sources:
        out.append(f'  {ts(s)},')
    out += [
        '];', '',
        '/** [문항ID, 민원, 정답, 오답후보, 해설, 난이도, 출처索引] */',
        'type Row = [string, string, string, string[], string, 1 | 2 | 3, number];', '',
        'const ROWS: Row[] = [',
    ]
    for q in rows:
        ds = '[' + ', '.join(ts(IDS[d][0]) for d in q['distractors']) + ']'
        out.append(f"  [{ts(q['id'])}, {ts(q['text'])}, {ts(IDS[q['correct']][0])}, {ds}, "
                   f"{ts(q['explanation'])}, {q['difficulty']}, {sources.index(q['source'])}],")
    out += [
        '];', '',
        'export const COMPLAINTS: Complaint[] = ROWS.map(',
        '  ([id, text, correctDept, distractors, explanation, difficulty, src]) => ({',
        '    id,',
        '    text,',
        '    correctDept,',
        '    distractors,',
        '    explanation,',
        '    difficulty,',
        '    source: SOURCES[src],',
        '  })',
        ');', '',
        'export const COMPLAINT_BY_ID = new Map(COMPLAINTS.map((c) => [c.id, c]));', '',
    ]
    (DATA / 'complaints.ts').write_text('\n'.join(out), encoding='utf-8')

    # ── 보고 ────────────────────────────────────────────────────────
    print(f'\n생성 완료 — 분류 {len(categories)}개 / 문항 {len(rows)}개')
    print(f'  제외: 미검수·보류 {skipped_review}건, 게임 미노출 분류 {skipped_cat}건')
    print('  난이도:', dict(sorted(Counter(q['difficulty'] for q in rows).items())))
    counts = Counter(q['correct'] for q in rows)
    print('  분류별:', ', '.join(f'{IDS[k][1]} {v}' for k, v in counts.items()))
    thin = [k for k, v in counts.items() if v < 10]
    if thin:
        print('  ⚠ 문항 10개 미만 분류:', ', '.join(thin))
    if len(rows) < 180:
        print(f'  ⚠ 문항 {len(rows)}개 — 1인 3회 플레이 시 중복이 체감됩니다(180개 이상 권장)')


if __name__ == '__main__':
    main()
